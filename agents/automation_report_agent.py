"""
Automation Report Agent
────────────────────────────────────────────────────────────
Generates Excel dashboard and PDF briefing from AutomationRiskAgent output.

Excel sheets:
  1. Risk Dashboard  — ranked table + bar chart
  2. Detailed Analysis — tasks, tech, recommendations per occupation
  3. Public Policy     — priority policies + monitoring indicators
  4. Resilient Skills  — occupations AI cannot easily replace

PDF:
  - KPI strip, ranking table, top-6 deep dives, policy section
"""

import pandas as pd
from pathlib import Path
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import letter
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table,
    TableStyle, HRFlowable, PageBreak, KeepTogether,
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT

OUTPUT_DIR = Path(__file__).parent.parent / "output"
OUTPUT_DIR.mkdir(exist_ok=True)

# ── Palette ───────────────────────────────────────────────────────────────────
NAVY   = "0F2240"
WHITE  = "FFFFFF"
LIGHT  = "F8FAFC"
DARK   = "1E293B"

RISK_HEX = {
    "CRITICAL": "DC2626",
    "HIGH":     "EA580C",
    "MEDIUM":   "CA8A04",
    "LOW":      "16A34A",
    "SAFE":     "2563EB",
}
RISK_RL = {k: colors.HexColor(f"#{v}") for k, v in RISK_HEX.items()}

# Page geometry
PAGE_W        = 8.5 * inch
MARGIN        = 0.65 * inch
CONTENT_WIDTH = PAGE_W - 2 * MARGIN   # 7.2 inches


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", start_color=hex_color, end_color=hex_color)

def _border() -> Border:
    s = Side(style="thin", color="E2E8F0")
    return Border(left=s, right=s, top=s, bottom=s)

def _cw(ws, col: int, w: float):
    ws.column_dimensions[get_column_letter(col)].width = w


# ════════════════════════════════════════════════════════════════════════════
#  EXCEL
# ════════════════════════════════════════════════════════════════════════════

class _ExcelBuilder:
    def __init__(self, wb: Workbook):
        self.wb = wb

    def _title(self, ws, text: str, subtitle: str = "", span: str = "A1:I1"):
        ws.merge_cells(span)
        c = ws[span.split(":")[0]]
        c.value, c.font, c.fill = (
            text,
            Font(bold=True, size=14, color=WHITE, name="Calibri"),
            _fill(NAVY),
        )
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30
        if subtitle:
            sub = span.replace("1", "2")
            ws.merge_cells(sub)
            s       = ws[sub.split(":")[0]]
            s.value = subtitle
            s.font  = Font(italic=True, size=10, color=DARK, name="Calibri")
            s.fill  = _fill(LIGHT)
            s.alignment = Alignment(horizontal="center")
            ws.row_dimensions[2].height = 16

    def _hdr(self, ws, row: int, cols: list, color: str = NAVY):
        for c, v in enumerate(cols, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.font      = Font(bold=True, color=WHITE, name="Calibri", size=9)
            cell.fill      = _fill(color)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = _border()
        ws.row_dimensions[row].height = 26

    def _row(self, ws, row: int, values: list, alt: bool = False, risk_col: int = None):
        bg = "F1F5F9" if alt else WHITE
        for c, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=c, value=v)
            if risk_col and c == risk_col and str(v) in RISK_HEX:
                cell.fill = _fill(RISK_HEX[str(v)])
                cell.font = Font(bold=True, color=WHITE, name="Calibri", size=9)
            else:
                cell.fill = _fill(bg)
                cell.font = Font(name="Calibri", size=9, color=DARK)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = _border()
        ws.row_dimensions[row].height = 22

    # ── Sheet 1: Risk Dashboard ──────────────────────────────────────────────

    def build_dashboard(self, results: list, metadata: dict):
        ws = self.wb.active
        ws.title = "Risk Dashboard"
        ws.sheet_view.showGridLines = False
        self._title(ws, "AI Automation Risk Index",
                    f"Analysis of {metadata['total_occupations']} occupations · "
                    f"Avg Score: {metadata['average_score']}/100")

        dist = metadata.get("distribution", {})
        kpis = [
            ("🔴 CRITICAL", dist.get("CRITICAL", 0), "DC2626"),
            ("🟠 HIGH",     dist.get("HIGH",     0), "EA580C"),
            ("🟡 MEDIUM",   dist.get("MEDIUM",   0), "CA8A04"),
            ("🟢 LOW",      dist.get("LOW",      0), "16A34A"),
            ("🔵 SAFE",     dist.get("SAFE",     0), "2563EB"),
        ]
        for i, (lbl, cnt, clr) in enumerate(kpis):
            col = i + 1
            for r_off, (val, sz, bold) in enumerate(
                [(lbl, 9, True), (str(cnt), 18, True), ("occupations", 8, False)]
            ):
                cell       = ws.cell(row=4 + r_off, column=col, value=val)
                cell.font  = Font(bold=bold, size=sz,
                                  color=WHITE if r_off < 2 else "DDDDDD", name="Calibri")
                cell.fill  = _fill(clr)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border    = _border()
            for r_h, h in [(4, 18), (5, 30), (6, 14)]:
                ws.row_dimensions[r_h].height = h

        headers = ["#", "Occupation", "SOC Code", "Score", "Risk Level",
                   "Horizon", "Routinization", "AI Capability", "Human Interact."]
        self._hdr(ws, 8, headers)

        sorted_r = sorted(results, key=lambda x: x["automation_score"], reverse=True)
        for i, r in enumerate(sorted_r):
            dim = r.get("dimensions", {})
            self._row(ws, 9 + i, [
                i + 1,
                r["title"],
                r["code"],
                f"{r['automation_score']:.0f}",
                r["risk_level"],
                r.get("time_horizon", "N/A"),
                f"{dim.get('routinization', 0):.0f}",
                f"{dim.get('ai_capability_current', 0):.0f}",
                f"{dim.get('human_interaction_required', 0):.0f}",
            ], alt=i % 2 == 0, risk_col=5)

        for c, w in enumerate([4, 38, 13, 8, 13, 22, 13, 14, 16], 1):
            _cw(ws, c, w)

        n = len(sorted_r)
        chart = BarChart()
        chart.type, chart.title = "bar", "Automation Risk Score by Occupation"
        chart.style, chart.height, chart.width = 10, 14, 24
        chart.add_data(Reference(ws, min_col=4, max_col=4,
                                 min_row=8, max_row=8 + n), titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=2, min_row=9, max_row=8 + n))
        ws.add_chart(chart, f"A{11 + n}")

    # ── Sheet 2: Detailed Analysis ───────────────────────────────────────────

    def build_details(self, results: list):
        ws = self.wb.create_sheet("Detailed Analysis")
        ws.sheet_view.showGridLines = False
        self._title(ws, "Detailed Analysis per Occupation",
                    "Tasks at risk · Threatening technologies · Recommendations",
                    "A1:H1")
        hdrs = ["Occupation", "Score", "Level", "Tasks at Risk",
                "Technologies", "Reasoning", "Worker Rec.", "Policy Rec."]
        self._hdr(ws, 4, hdrs)

        for i, r in enumerate(sorted(results,
                                     key=lambda x: x["automation_score"], reverse=True)):
            self._row(ws, 5 + i, [
                r["title"],
                f"{r['automation_score']:.0f}/100",
                r["risk_level"],
                "; ".join(r.get("tasks_at_risk", [])[:3]),
                "; ".join(r.get("threatening_technologies", [])[:3]),
                r.get("reasoning", ""),
                r.get("worker_recommendation", ""),
                r.get("policy_recommendation", ""),
            ], alt=i % 2 == 0, risk_col=3)
            ws.row_dimensions[5 + i].height = 38

        for c, w in enumerate([32, 9, 12, 36, 30, 40, 36, 36], 1):
            _cw(ws, c, w)

    # ── Sheet 3: Policy ──────────────────────────────────────────────────────

    def build_policy(self, policy: dict, metadata: dict):
        ws = self.wb.create_sheet("Public Policy")
        ws.sheet_view.showGridLines = False
        self._title(ws, "Public Policy Recommendations",
                    f"Urgency: {policy.get('policy_urgency','N/A')} · "
                    f"{metadata['total_occupations']} occupations analyzed",
                    "A1:F1")

        ws.merge_cells("A3:F3")
        ws["A3"].value = "OVERALL DIAGNOSIS"
        ws["A3"].font  = Font(bold=True, color=WHITE, name="Calibri", size=10)
        ws["A3"].fill  = _fill(NAVY)
        ws["A3"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A4:F6")
        ws["A4"].value = policy.get("overall_diagnosis", "")
        ws["A4"].font  = Font(name="Calibri", size=10, color=DARK)
        ws["A4"].fill  = _fill(LIGHT)
        ws["A4"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[4].height = 50

        self._hdr(ws, 8, ["#", "Policy", "Description", "Target Population", "Horizon"])
        for i, p in enumerate(policy.get("priority_policies", [])):
            self._row(ws, 9 + i, [
                i + 1, p.get("name", ""), p.get("description", ""),
                p.get("target_population", ""), p.get("horizon", ""),
            ], alt=i % 2 == 0)
            ws.row_dimensions[9 + i].height = 30

        pol_n   = len(policy.get("priority_policies", []))
        ind_row = 11 + pol_n
        ws.merge_cells(f"A{ind_row}:F{ind_row}")
        ws[f"A{ind_row}"].value = "MONITORING INDICATORS"
        ws[f"A{ind_row}"].font  = Font(bold=True, color=WHITE, name="Calibri", size=10)
        ws[f"A{ind_row}"].fill  = _fill("2A9D8F")
        ws[f"A{ind_row}"].alignment = Alignment(horizontal="center")

        for j, ind in enumerate(policy.get("monitoring_indicators", [])):
            ws.merge_cells(f"A{ind_row+1+j}:F{ind_row+1+j}")
            cell       = ws[f"A{ind_row+1+j}"]
            cell.value = f"  {j+1}. {ind}"
            cell.font  = Font(name="Calibri", size=10, color=DARK)
            cell.fill  = _fill(LIGHT if j % 2 == 0 else WHITE)

        msg_row = ind_row + len(policy.get("monitoring_indicators", [])) + 3
        ws.merge_cells(f"A{msg_row}:F{msg_row}")
        ws[f"A{msg_row}"].value = f"💡 {policy.get('key_message', '')}"
        ws[f"A{msg_row}"].font  = Font(bold=True, name="Calibri", size=11, color=WHITE)
        ws[f"A{msg_row}"].fill  = _fill(NAVY)
        ws[f"A{msg_row}"].alignment = Alignment(horizontal="center", wrap_text=True)
        ws.row_dimensions[msg_row].height = 28

        for c, w in enumerate([4, 28, 42, 28, 18], 1):
            _cw(ws, c, w)

    # ── Sheet 4: Resilient Skills ────────────────────────────────────────────

    def build_resilience(self, results: list):
        ws = self.wb.create_sheet("Resilient Skills")
        ws.sheet_view.showGridLines = False
        self._title(ws, "Resilient Occupations — What AI Cannot Replace",
                    "Low automation risk · High human-value skills", "A1:E1")

        self._hdr(ws, 4, ["Occupation", "Score", "Level", "Resilient Tasks", "Worker Action"])
        low_risk = sorted([r for r in results if r["automation_score"] < 50],
                          key=lambda x: x["automation_score"])
        for i, r in enumerate(low_risk):
            self._row(ws, 5 + i, [
                r["title"],
                f"{r['automation_score']:.0f}/100",
                r["risk_level"],
                "; ".join(r.get("resilient_tasks", [])[:2]),
                r.get("worker_recommendation", ""),
            ], alt=i % 2 == 0, risk_col=3)
            ws.row_dimensions[5 + i].height = 28

        for c, w in enumerate([34, 10, 12, 40, 42], 1):
            _cw(ws, c, w)


# ════════════════════════════════════════════════════════════════════════════
#  PDF
# ════════════════════════════════════════════════════════════════════════════

class _PDFBuilder:
    def __init__(self, path: Path):
        self.path = path
        self.S    = getSampleStyleSheet()
        self._styles()

    def _styles(self):
        S = self.S
        def add(name, **kw):
            if name not in S:
                S.add(ParagraphStyle(name, **kw))

        add("ARTitle",
            parent=S["Title"], fontSize=20,
            textColor=colors.HexColor("#0F2240"),
            fontName="Helvetica-Bold", spaceAfter=4)
        add("ARSection",
            parent=S["Heading1"], fontSize=12,
            textColor=colors.HexColor("#0F2240"),
            fontName="Helvetica-Bold", spaceBefore=12, spaceAfter=4)
        add("ARBody",
            parent=S["Normal"], fontSize=9.5,
            leading=14, textColor=colors.HexColor("#1E293B"), spaceAfter=6)
        add("ARCaption",
            parent=S["Normal"], fontSize=8,
            textColor=colors.grey, alignment=TA_CENTER)
        add("ARCell",
            parent=S["Normal"], fontSize=8.5,
            leading=12, textColor=colors.HexColor("#1E293B"))
        add("ARCellBold",
            parent=S["Normal"], fontSize=8.5,
            leading=12, fontName="Helvetica-Bold",
            textColor=colors.HexColor("#1E293B"))
        add("ARCellWhite",
            parent=S["Normal"], fontSize=9,
            leading=13, textColor=colors.white)
        add("ARCellWhiteBold",
            parent=S["Normal"], fontSize=9,
            leading=13, fontName="Helvetica-Bold", textColor=colors.white)

    def _hr(self, color="#0F2240", thickness=1.5):
        return HRFlowable(width="100%", thickness=thickness,
                          color=colors.HexColor(color), spaceAfter=6)

    def _p(self, text: str, style = "ARCell") -> Paragraph:
        if isinstance(style, str):
            style = self.S[style]
        return Paragraph(str(text) if text else "", style)

    def _ranking_table(self, rows_data: list) -> Table:
        """Ranking table — all cells wrapped in Paragraphs."""
        # Column widths: # | Occupation | Level | Score | Horizon
        widths = [0.3*inch, 2.9*inch, 0.85*inch, 0.55*inch, 2.6*inch]

        header = [
            self._p("＃",          "ARCellWhiteBold"),
            self._p("Occupation",  "ARCellWhiteBold"),
            self._p("Level",       "ARCellWhiteBold"),
            self._p("Score",       "ARCellWhiteBold"),
            self._p("Time Horizon","ARCellWhiteBold"),
        ]
        table_rows = [header]

        risk_colors = {}
        for i, (num, title, level, score, horizon) in enumerate(rows_data):
            row = [
                self._p(str(num)),
                self._p(title),
                self._p(level),
                self._p(str(score)),
                self._p(horizon),
            ]
            table_rows.append(row)
            risk_colors[i + 1] = RISK_RL.get(level, colors.grey)

        t = Table(table_rows, colWidths=widths)
        style_cmds = [
            ("BACKGROUND",    (0, 0), (-1,  0), colors.HexColor("#0F2240")),
            ("GRID",          (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
            ("ROWBACKGROUNDS",(0, 1), (-1, -1),
             [colors.HexColor("#F8FAFC"), colors.white]),
            ("TOPPADDING",    (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING",   (0, 0), (-1, -1), 5),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 5),
            ("VALIGN",        (0, 0), (-1, -1), "TOP"),
        ]
        for row_idx, clr in risk_colors.items():
            style_cmds += [
                ("BACKGROUND", (2, row_idx), (2, row_idx), clr),
                ("TEXTCOLOR",  (2, row_idx), (2, row_idx), colors.white),
            ]
        t.setStyle(TableStyle(style_cmds))
        return t

    def _policy_table(self, policies: list) -> Table:
        """Policy table — all cells wrapped in Paragraphs, proper widths."""
        # Total = 7.2 inches: Policy | Description | Target Population | Horizon
        widths = [1.6*inch, 3.2*inch, 1.5*inch, 0.9*inch]

        header = [
            self._p("Policy",            "ARCellWhiteBold"),
            self._p("Description",       "ARCellWhiteBold"),
            self._p("Target Population", "ARCellWhiteBold"),
            self._p("Horizon",           "ARCellWhiteBold"),
        ]
        table_rows = [header]
        for p in policies:
            table_rows.append([
                self._p(p.get("name", "")),
                self._p(p.get("description", "")),
                self._p(p.get("target_population", "")),
                self._p(p.get("horizon", "")),
            ])

        t = Table(table_rows, colWidths=widths)
        t.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1,  0), colors.HexColor("#0F2240")),
            ("GRID",          (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
            ("ROWBACKGROUNDS",(0, 1), (-1, -1),
             [colors.HexColor("#F8FAFC"), colors.white]),
            ("TOPPADDING",    (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LEFTPADDING",   (0, 0), (-1, -1), 6),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 6),
            ("VALIGN",        (0, 0), (-1, -1), "TOP"),
            ("LINEBEFORE",    (0, 0), (0, -1),  3, colors.HexColor("#DC2626")),
        ]))
        return t

    def _deep_dive_card(self, r: dict) -> list:
        """Single occupation deep-dive card — all cells Paragraphs."""
        clr = RISK_RL.get(r["risk_level"], colors.grey)

        # Header bar: title left, score/level right
        hdr = Table(
            [[
                self._p(r["title"], "ARCellWhiteBold"),
                self._p(f"{r['automation_score']:.0f}/100 — {r['risk_level']}",
                        ParagraphStyle("hs", parent=self.S["ARCellWhiteBold"],
                                       alignment=TA_RIGHT)),
            ]],
            colWidths=[4.5 * inch, 2.7 * inch],
        )
        hdr.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, -1), clr),
            ("TOPPADDING",    (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("LEFTPADDING",   (0, 0), (0,  -1), 8),
            ("RIGHTPADDING",  (-1, 0),(-1, -1), 8),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ]))

        # Detail rows — label col narrow, content col fills the rest
        label_w   = 1.15 * inch
        content_w = CONTENT_WIDTH - label_w

        def row(label, text):
            return [self._p(label, "ARCellBold"), self._p(text, "ARCell")]

        detail = Table(
            [
                row("Reasoning",     r.get("reasoning", "")),
                row("Tasks at Risk", "; ".join(r.get("tasks_at_risk",  [])[:3])),
                row("Technologies",  "; ".join(r.get("threatening_technologies", [])[:3])),
                row("Worker Action", r.get("worker_recommendation", "")),
                row("Policy Action", r.get("policy_recommendation", "")),
            ],
            colWidths=[label_w, content_w],
        )
        detail.setStyle(TableStyle([
            ("GRID",          (0, 0), (-1, -1), 0.3, colors.HexColor("#E2E8F0")),
            ("ROWBACKGROUNDS",(0, 0), (-1, -1),
             [colors.HexColor("#F8FAFC"), colors.white]),
            ("TOPPADDING",    (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING",   (0, 0), (-1, -1), 6),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 6),
            ("VALIGN",        (0, 0), (-1, -1), "TOP"),
        ]))

        return [hdr, detail, Spacer(1, 8)]

    def build(self, analysis: dict):
        results  = analysis["results"]
        policy   = analysis["policy_report"]
        metadata = analysis["metadata"]
        S        = self.S

        doc = SimpleDocTemplate(
            str(self.path), pagesize=letter,
            rightMargin=MARGIN, leftMargin=MARGIN,
            topMargin=0.65 * inch, bottomMargin=0.65 * inch,
        )
        story = []

        # ── Title ─────────────────────────────────────────────────────────────
        story += [
            Paragraph("AI Automation Risk Index", S["ARTitle"]),
            Paragraph(
                f"Analysis of {metadata['total_occupations']} occupations &nbsp;|&nbsp; "
                f"Avg Score: <b>{metadata['average_score']}/100</b> &nbsp;|&nbsp; "
                f"{datetime.now().strftime('%B %Y')}",
                ParagraphStyle("sub", parent=S["ARBody"],
                               textColor=colors.HexColor("#64748B"), fontSize=10),
            ),
            self._hr(),
        ]

        # ── KPI strip ─────────────────────────────────────────────────────────
        dist     = metadata.get("distribution", {})
        kpi_keys = ["CRITICAL", "HIGH", "MEDIUM", "LOW", "SAFE"]
        kpi_clrs = [colors.HexColor(f"#{RISK_HEX[k]}") for k in kpi_keys]
        col_w    = CONTENT_WIDTH / len(kpi_keys)

        kpi_top = [self._p(k, ParagraphStyle(
            f"kl_{k}", parent=S["Normal"], fontSize=8,
            fontName="Helvetica-Bold", textColor=colors.white, alignment=TA_CENTER,
        )) for k in kpi_keys]

        kpi_bot = [self._p(str(dist.get(k, 0)), ParagraphStyle(
            f"kv_{k}", parent=S["Normal"], fontSize=20,
            fontName="Helvetica-Bold", textColor=colors.white, alignment=TA_CENTER,
        )) for k in kpi_keys]

        kpi_sub = [self._p("occupations", ParagraphStyle(
            f"ks_{k}", parent=S["Normal"], fontSize=7,
            textColor=colors.HexColor("#DDDDDD"), alignment=TA_CENTER,
        )) for k in kpi_keys]

        kpi_t = Table(
            [kpi_top, kpi_bot, kpi_sub],
            colWidths=[col_w] * len(kpi_keys),
        )
        kpi_style = [
            ("TOPPADDING",    (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]
        for ci, clr in enumerate(kpi_clrs):
            kpi_style.append(("BACKGROUND", (ci, 0), (ci, -1), clr))
        kpi_t.setStyle(TableStyle(kpi_style))
        story += [kpi_t, Spacer(1, 14)]

        # ── Diagnosis ─────────────────────────────────────────────────────────
        story += [
            Paragraph("Overall Diagnosis", S["ARSection"]),
            Paragraph(policy.get("overall_diagnosis", ""), S["ARBody"]),
            Spacer(1, 6),
        ]

        # ── Ranking table ─────────────────────────────────────────────────────
        story += [self._hr(), Paragraph("Risk Ranking by Occupation", S["ARSection"])]
        sorted_r = sorted(results, key=lambda x: x["automation_score"], reverse=True)
        rows_data = [
            (i + 1, r["title"], r["risk_level"],
             f"{r['automation_score']:.0f}", r.get("time_horizon", "N/A"))
            for i, r in enumerate(sorted_r)
        ]
        story += [self._ranking_table(rows_data), Spacer(1, 14)]

        # ── Deep dives (top 6) ────────────────────────────────────────────────
        story.append(PageBreak())
        story.append(Paragraph("Detailed Analysis — High Exposure", S["ARSection"]))
        story.append(Spacer(1, 6))
        for r in sorted_r[:6]:
            story += self._deep_dive_card(r)

        # ── Policy section ────────────────────────────────────────────────────
        story += [
            PageBreak(),
            self._hr(),
            Paragraph("Public Policy Recommendations", S["ARSection"]),
            Spacer(1, 6),
            self._policy_table(policy.get("priority_policies", [])),
            Spacer(1, 14),
        ]

        # Key message box
        msg = Table(
            [[self._p(
                f"&#9632; {policy.get('key_message', '')}",
                ParagraphStyle("msg", parent=S["Normal"], fontSize=11,
                               fontName="Helvetica-Bold",
                               textColor=colors.white, alignment=TA_LEFT),
            )]],
            colWidths=[CONTENT_WIDTH],
        )
        msg.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, -1), colors.HexColor("#0F2240")),
            ("TOPPADDING",    (0, 0), (-1, -1), 12),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
            ("LEFTPADDING",   (0, 0), (-1, -1), 14),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 14),
            ("LINEBEFORE",    (0, 0), (0, -1),  4, colors.HexColor("#DC2626")),
        ]))
        story += [msg, Spacer(1, 20)]

        # ── Footer ────────────────────────────────────────────────────────────
        story += [
            HRFlowable(width="100%", thickness=0.5,
                       color=colors.HexColor("#CBD5E1")),
            Paragraph(
                "Generated by LABORFLEX — Automation Risk Agent | "
                "AI Analysis: Claude | Occupation Data: O*NET Web Services",
                ParagraphStyle("ftr", parent=S["Normal"], fontSize=7,
                               textColor=colors.grey, alignment=TA_CENTER),
            ),
        ]

        doc.build(story)
        print(f"[AutomationReportAgent] PDF saved: {self.path}")


# ════════════════════════════════════════════════════════════════════════════
#  Public Agent
# ════════════════════════════════════════════════════════════════════════════

class AutomationReportAgent:
    def generate(self, analysis: dict) -> dict:
        """
        Generate Excel + PDF reports from AutomationRiskAgent output.
        Returns {"excel": Path, "pdf": Path}.
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")

        xl_path = OUTPUT_DIR / f"automation_risk_{timestamp}.xlsx"
        wb      = Workbook()
        b       = _ExcelBuilder(wb)
        b.build_dashboard(analysis["results"], analysis["metadata"])
        b.build_details(analysis["results"])
        b.build_policy(analysis["policy_report"], analysis["metadata"])
        b.build_resilience(analysis["results"])
        wb.save(xl_path)
        print(f"[AutomationReportAgent] Excel saved: {xl_path}")

        pdf_path = OUTPUT_DIR / f"automation_risk_{timestamp}.pdf"
        _PDFBuilder(pdf_path).build(analysis)

        return {"excel": xl_path, "pdf": pdf_path}