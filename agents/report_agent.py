"""
Report Agent
────────────────────────────────────────────────────────────
Generates Excel dashboard and PDF executive briefing from
EconomicAnalysisAgent output (general labor market analysis).

Excel sheets:
  1. Executive Summary
  2. Unemployment Data
  3. Wage Growth
  4. Occupation Profiles (O*NET)
  5. Risk Factors

PDF:
  - KPI table, unemployment insights, wage dynamics,
    structural analysis, O*NET occupations, policy footer
"""

import pandas as pd
from pathlib import Path
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import letter
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable, KeepTogether
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch

OUTPUT_DIR = Path(__file__).parent.parent / "output"
OUTPUT_DIR.mkdir(exist_ok=True)

NAVY  = "1B3A5C"
TEAL  = "2A9D8F"
GOLD  = "E9C46A"
LIGHT = "F4F6F8"
WHITE = "FFFFFF"
DARK  = "2D3748"


def _fill(h):  return PatternFill("solid", start_color=h, end_color=h)
def _border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)
def _cw(ws, col, w):
    ws.column_dimensions[get_column_letter(col)].width = w


class _ExcelBuilder:
    def __init__(self, wb):
        self.wb = wb

    def _title(self, ws, text, subtitle="", span="A1:H1"):
        ws.merge_cells(span)
        c       = ws[span.split(":")[0]]
        c.value = text
        c.font  = Font(bold=True, size=14, color=WHITE, name="Arial")
        c.fill  = _fill(NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28
        if subtitle:
            s2 = span.replace("1","2")
            ws.merge_cells(s2)
            s       = ws[s2.split(":")[0]]
            s.value = subtitle
            s.font  = Font(italic=True, size=10, color=DARK, name="Arial")
            s.fill  = _fill(LIGHT)
            s.alignment = Alignment(horizontal="center")
            ws.row_dimensions[2].height = 18

    def _hdr(self, ws, row, cols, color=NAVY):
        for c, v in enumerate(cols, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.font      = Font(bold=True, color=WHITE, name="Arial", size=10)
            cell.fill      = _fill(color)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = _border()
        ws.row_dimensions[row].height = 28

    def _row(self, ws, row, values, alt=False):
        bg = _fill("F0F4F8") if alt else _fill(WHITE)
        for c, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.font      = Font(name="Arial", size=9, color=DARK)
            cell.fill      = bg
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border    = _border()
        ws.row_dimensions[row].height = 20

    def build_summary(self, insights: dict):
        ws = self.wb.active
        ws.title = "Executive Summary"
        ws.sheet_view.showGridLines = False
        self._title(ws, "Labor Market Intelligence Report",
                    f"Report Date: {insights.get('report_date', datetime.now().strftime('%B %Y'))}")

        ws.merge_cells("A4:H12")
        c       = ws["A4"]
        c.value = insights.get("executive_summary", "No summary available.")
        c.font  = Font(name="Arial", size=10, color=DARK)
        c.fill  = _fill(LIGHT)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        for r in range(4,13): ws.row_dimensions[r].height = 16

        row = 14
        ws.merge_cells(f"A{row}:H{row}")
        h       = ws[f"A{row}"]
        h.value = "KEY INDICATORS"
        h.font  = Font(bold=True, size=11, color=WHITE, name="Arial")
        h.fill  = _fill(TEAL)
        h.alignment = Alignment(horizontal="center")
        ws.row_dimensions[row].height = 22

        unemp  = insights.get("unemployment", {})
        wages  = insights.get("wages", {})
        policy = insights.get("policy_structural", {})
        kpis   = [
            ("Trend Direction",   unemp.get("trend_direction","N/A")),
            ("Headline Signal",   unemp.get("headline","N/A")),
            ("Wage Trend",        wages.get("overall_wage_trend","N/A")),
            ("Wage Dispersion",   wages.get("wage_dispersion_signal","N/A")),
            ("Structural Shift",  policy.get("structural_shift","N/A")),
            ("Policy Focus",      policy.get("recommended_policy_focus","N/A")),
        ]
        self._hdr(ws, row+1, ["Indicator","Finding"], NAVY)
        for i,(lbl,val) in enumerate(kpis):
            self._row(ws, row+2+i, [lbl, val], alt=i%2==0)
            ws.merge_cells(f"B{row+2+i}:H{row+2+i}")

        _cw(ws, 1, 28)
        for c in range(2,9): _cw(ws, c, 18)

    def build_unemployment(self, csv_datasets, insights):
        ws = self.wb.create_sheet("Unemployment Data")
        ws.sheet_view.showGridLines = False
        self._title(ws, "Unemployment & Labor Force Participation", "Monthly Trend Data")

        df = csv_datasets.get("unemployment", pd.DataFrame())
        if df.empty:
            ws["A4"] = "No unemployment.csv found in /data."
            return

        cols = list(df.columns)
        self._hdr(ws, 4, cols)
        for i, row_data in enumerate(df.itertuples(index=False)):
            self._row(ws, 5+i, list(row_data), alt=i%2==0)

        unemp = insights.get("unemployment", {})
        br    = 5 + len(df) + 2
        ws.merge_cells(f"A{br}:H{br}")
        ws[f"A{br}"].value = "🔍 AI INSIGHT"
        ws[f"A{br}"].font  = Font(bold=True, color=WHITE, name="Arial")
        ws[f"A{br}"].fill  = _fill(TEAL)
        ws[f"A{br}"].alignment = Alignment(horizontal="center")
        ws.merge_cells(f"A{br+1}:H{br+3}")
        ws[f"A{br+1}"].value = (
            f"{unemp.get('key_finding','')}\n\n"
            f"Phillips Curve Signal: {unemp.get('phillips_curve_signal','')}"
        )
        ws[f"A{br+1}"].font      = Font(name="Arial", size=10, color=DARK)
        ws[f"A{br+1}"].fill      = _fill(LIGHT)
        ws[f"A{br+1}"].alignment = Alignment(wrap_text=True, vertical="top")

        for c in range(1, len(cols)+1): _cw(ws, c, 22)

        if df.select_dtypes(include="number").shape[1] > 0:
            chart = LineChart()
            chart.title  = "Unemployment Trend"
            chart.style  = 10
            chart.height = 10
            chart.width  = 20
            n = len(df)
            chart.add_data(Reference(ws, min_col=2, max_col=min(3,len(cols)),
                                     min_row=4, max_row=4+n), titles_from_data=True)
            ws.add_chart(chart, f"A{br+5}")

    def build_wages(self, csv_datasets, insights):
        ws = self.wb.create_sheet("Wage Growth")
        ws.sheet_view.showGridLines = False
        self._title(ws, "Wage Growth by Sector", "Average Hourly Wage & YoY Growth %")

        df = csv_datasets.get("wage_growth", pd.DataFrame())
        if df.empty:
            ws["A4"] = "No wage_growth.csv found in /data."
            return

        cols = list(df.columns)
        self._hdr(ws, 4, cols)
        for i, row_data in enumerate(df.itertuples(index=False)):
            self._row(ws, 5+i, list(row_data), alt=i%2==0)

        w    = insights.get("wages", {})
        br   = 5 + len(df) + 2
        ws.merge_cells(f"A{br}:H{br}")
        ws[f"A{br}"].value = "🔍 AI WAGE INSIGHT"
        ws[f"A{br}"].font  = Font(bold=True, color=WHITE, name="Arial")
        ws[f"A{br}"].fill  = _fill(TEAL)
        ws[f"A{br}"].alignment = Alignment(horizontal="center")
        ws.merge_cells(f"A{br+1}:H{br+3}")
        ws[f"A{br+1}"].value = (
            f"Sector Leaders: {', '.join(w.get('sector_leaders',[]))}\n"
            f"Real Wage Assessment: {w.get('real_wage_assessment','')}\n"
            f"Policy Implication: {w.get('policy_implication','')}"
        )
        ws[f"A{br+1}"].font      = Font(name="Arial", size=10, color=DARK)
        ws[f"A{br+1}"].fill      = _fill(LIGHT)
        ws[f"A{br+1}"].alignment = Alignment(wrap_text=True, vertical="top")

        for c in range(1, len(cols)+1): _cw(ws, c, 22)

    def build_onet(self, insights):
        ws = self.wb.create_sheet("Occupation Profiles")
        ws.sheet_view.showGridLines = False
        self._title(ws, "O*NET Occupation Profiles", "High-Demand Labor Market Roles")

        records = insights.get("onet_occupations", [])
        if not records:
            ws["A4"] = "No O*NET data collected."
            return

        df   = pd.DataFrame(records)
        cols = list(df.columns)
        self._hdr(ws, 4, [c.title() for c in cols])
        for i, row_data in enumerate(df.itertuples(index=False)):
            self._row(ws, 5+i, list(row_data), alt=i%2==0)

        p   = insights.get("policy_structural", {})
        br  = 5 + len(df) + 2
        ws.merge_cells(f"A{br}:H{br}")
        ws[f"A{br}"].value = "🔍 STRUCTURAL ANALYSIS"
        ws[f"A{br}"].font  = Font(bold=True, color=WHITE, name="Arial")
        ws[f"A{br}"].fill  = _fill(NAVY)
        ws[f"A{br}"].alignment = Alignment(horizontal="center")
        ws.merge_cells(f"A{br+1}:H{br+3}")
        ws[f"A{br+1}"].value = (
            f"High-Demand Roles: {', '.join(p.get('high_demand_occupations',[]))}\n"
            f"Structural Shift: {p.get('structural_shift','')}\n"
            f"Automation Risk Signal: {p.get('automation_risk_signal','')}"
        )
        ws[f"A{br+1}"].font      = Font(name="Arial", size=10, color=DARK)
        ws[f"A{br+1}"].fill      = _fill(LIGHT)
        ws[f"A{br+1}"].alignment = Alignment(wrap_text=True, vertical="top")

        for c in range(1, len(cols)+1): _cw(ws, c, 24)

    def build_risks(self, insights):
        ws = self.wb.create_sheet("Risk Factors")
        ws.sheet_view.showGridLines = False
        self._title(ws, "Labor Market Risk Assessment", "AI-Generated Risk Signals")

        risks = insights.get("unemployment", {}).get("risk_factors", [])
        self._hdr(ws, 4, ["#","Risk Factor","Category"])
        cats  = ["Structural","Cyclical","Policy"]
        for i, risk in enumerate(risks):
            self._row(ws, 5+i, [i+1, risk, cats[i % len(cats)]], alt=i%2==0)

        conf    = insights.get("policy_structural",{}).get("confidence_level","N/A")
        nr      = 7 + len(risks)
        ws.merge_cells(f"A{nr}:H{nr}")
        ws[f"A{nr}"].value = f"Analysis Confidence Level: {conf}"
        ws[f"A{nr}"].font  = Font(bold=True, name="Arial", color=DARK)
        ws[f"A{nr}"].fill  = _fill(GOLD)
        ws[f"A{nr}"].alignment = Alignment(horizontal="center")

        for c, w in [(1,6),(2,55),(3,16)]: _cw(ws, c, w)


class _PDFBuilder:
    def __init__(self, path):
        self.path = path
        self.S    = getSampleStyleSheet()
        self._add_styles()

    def _add_styles(self):
        self.S.add(ParagraphStyle("LMTitle",
            parent=self.S["Title"], fontSize=26,
            textColor=colors.white, fontName="Helvetica-Bold",
            spaceAfter=0, alignment=1))
        self.S.add(ParagraphStyle("LMSubtitle",
            parent=self.S["Normal"], fontSize=11,
            textColor=colors.HexColor("#A8C4D8"), fontName="Helvetica-Oblique",
            spaceAfter=0, alignment=1))
        self.S.add(ParagraphStyle("LMSection",
            parent=self.S["Normal"], fontSize=11,
            textColor=colors.white, fontName="Helvetica-Bold",
            spaceBefore=0, spaceAfter=0))
        self.S.add(ParagraphStyle("LMBody",
            parent=self.S["Normal"], fontSize=10,
            leading=16, textColor=colors.HexColor("#2D3748"), spaceAfter=6))
        self.S.add(ParagraphStyle("LMLabel",
            parent=self.S["Normal"], fontSize=9,
            textColor=colors.HexColor("#2A9D8F"), fontName="Helvetica-Bold",
            spaceAfter=2))

    def _header_block(self, title, subtitle):
        data = [
            [Paragraph(title,    self.S["LMTitle"])],
            [Paragraph(subtitle, self.S["LMSubtitle"])],
        ]
        t = Table(data, colWidths=[7.0 * inch])
        t.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, -1), colors.HexColor("#1B3A5C")),
            ("TOPPADDING",    (0, 0), (0,  0),  22),
            ("BOTTOMPADDING", (0, 0), (0,  0),  4),
            ("TOPPADDING",    (0, 1), (0,  1),  4),
            ("BOTTOMPADDING", (0, 1), (0,  1),  22),
            ("LEFTPADDING",   (0, 0), (-1, -1), 20),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 20),
        ]))
        return t

    def _section_header(self, text):
        data = [[Paragraph(text, self.S["LMSection"])]]
        t = Table(data, colWidths=[7.0 * inch])
        t.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, -1), colors.HexColor("#2A9D8F")),
            ("TOPPADDING",    (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ("LEFTPADDING",   (0, 0), (-1, -1), 12),
        ]))
        return t

    def _kv_table(self, rows, widths=None) -> Table:
        th = ParagraphStyle("th", parent=self.S["Normal"],
            fontSize=9, fontName="Helvetica-Bold", textColor=colors.white)
        td_lbl = ParagraphStyle("td_lbl", parent=self.S["Normal"],
            fontSize=9, fontName="Helvetica-Bold",
            textColor=colors.HexColor("#1B3A5C"))
        td_val = ParagraphStyle("td_val", parent=self.S["Normal"],
            fontSize=9, leading=14, textColor=colors.HexColor("#2D3748"))

        styled = []
        for i, row in enumerate(rows):
            if i == 0:
                styled.append([Paragraph(str(v), th) for v in row])
            else:
                styled.append([Paragraph(str(row[0]), td_lbl),
                                Paragraph(str(row[1]), td_val)])

        t = Table(styled, colWidths=widths)
        t.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1,  0), colors.HexColor("#1B3A5C")),
            ("LINEBELOW",     (0, 0), (-1,  0), 1.5, colors.HexColor("#2A9D8F")),
            ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.HexColor("#EBF4FB"),
                                                  colors.white]),
            ("GRID",          (0, 0), (-1, -1), 0.5, colors.HexColor("#D1D9E0")),
            ("TOPPADDING",    (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ("LEFTPADDING",   (0, 0), (-1, -1), 10),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 10),
            ("VALIGN",        (0, 0), (-1, -1), "TOP"),
        ]))
        return t

    def _kpi_grid(self, kpis):
        lbl_s = ParagraphStyle("kpi_lbl", parent=self.S["Normal"],
            fontSize=8, fontName="Helvetica-Bold",
            textColor=colors.HexColor("#2A9D8F"), spaceAfter=3)
        val_s = ParagraphStyle("kpi_val", parent=self.S["Normal"],
            fontSize=10, fontName="Helvetica-Bold",
            textColor=colors.HexColor("#1B3A5C"), leading=13)

        cells = [[Paragraph(lbl.upper(), lbl_s), Paragraph(str(val), val_s)]
                 for lbl, val in kpis]
        # Wrap into 2-column card rows
        card_cells = [[Paragraph(lbl.upper(), lbl_s),
                       Paragraph(str(val), val_s)]
                      for lbl, val in kpis]
        rows = []
        for i in range(0, len(kpis), 2):
            left  = [Paragraph(kpis[i][0].upper(),   lbl_s),
                     Paragraph(str(kpis[i][1]),       val_s)]
            right = ([Paragraph(kpis[i+1][0].upper(), lbl_s),
                      Paragraph(str(kpis[i+1][1]),    val_s)]
                     if i + 1 < len(kpis) else ["", ""])
            rows.append([left, right])

        col_w = 3.35 * inch
        t = Table(rows, colWidths=[col_w, col_w])
        t.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, -1), colors.HexColor("#EBF4FB")),
            ("BOX",           (0, 0), (-1, -1), 0.5, colors.HexColor("#D1D9E0")),
            ("INNERGRID",     (0, 0), (-1, -1), 0.5, colors.HexColor("#D1D9E0")),
            ("TOPPADDING",    (0, 0), (-1, -1), 10),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
            ("LEFTPADDING",   (0, 0), (-1, -1), 12),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 12),
            ("VALIGN",        (0, 0), (-1, -1), "TOP"),
        ]))
        return t

    def _risks_table(self, risks):
        cats       = ["Structural", "Cyclical", "Policy"]
        cat_colors = {"Structural": "#C53030", "Cyclical": "#C05621", "Policy": "#2B6CB0"}
        th_s  = ParagraphStyle("rth", parent=self.S["Normal"],
            fontSize=9, fontName="Helvetica-Bold", textColor=colors.white)
        row_s = ParagraphStyle("rtd", parent=self.S["Normal"],
            fontSize=9, leading=14, textColor=colors.HexColor("#2D3748"))
        cat_s = ParagraphStyle("rtc", parent=self.S["Normal"],
            fontSize=8, fontName="Helvetica-Bold", textColor=colors.white, alignment=1)

        rows = [[Paragraph("#", th_s), Paragraph("Risk Factor", th_s),
                 Paragraph("Category", th_s)]]
        style_cmds = [
            ("BACKGROUND",    (0, 0), (-1, 0),  colors.HexColor("#1B3A5C")),
            ("LINEBELOW",     (0, 0), (-1, 0),  1.5, colors.HexColor("#2A9D8F")),
            ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.HexColor("#FFF5F5"),
                                                  colors.white]),
            ("GRID",          (0, 0), (-1, -1), 0.5, colors.HexColor("#D1D9E0")),
            ("TOPPADDING",    (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("LEFTPADDING",   (0, 0), (-1, -1), 8),
            ("VALIGN",        (0, 0), (-1, -1), "TOP"),
            ("ALIGN",         (0, 0), (0,  -1), "CENTER"),
        ]
        for i, risk in enumerate(risks):
            cat = cats[i % len(cats)]
            rows.append([Paragraph(str(i + 1), row_s),
                         Paragraph(risk, row_s),
                         Paragraph(cat, cat_s)])
            style_cmds.append(("BACKGROUND", (2, i + 1), (2, i + 1),
                                colors.HexColor(cat_colors[cat])))

        t = Table(rows, colWidths=[0.4 * inch, 5.4 * inch, 1.2 * inch])
        t.setStyle(TableStyle(style_cmds))
        return t

    def _page_footer(self, canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.HexColor("#718096"))
        footer_text = (
            "Generated by LABORFLEX Labor Market Intelligence Agent  |  "
            "Powered by Claude AI  |  Data: O*NET Web Services + BLS OEWS"
        )
        canvas.drawCentredString(letter[0] / 2, 0.45 * inch, footer_text)
        canvas.drawRightString(letter[0] - 0.75 * inch, 0.45 * inch,
                               f"Page {doc.page}")
        canvas.setStrokeColor(colors.HexColor("#CCCCCC"))
        canvas.setLineWidth(0.5)
        canvas.line(0.75 * inch, 0.58 * inch,
                    letter[0] - 0.75 * inch, 0.58 * inch)
        canvas.restoreState()

    def build(self, insights, csv_datasets):
        doc = SimpleDocTemplate(
            str(self.path), pagesize=letter,
            rightMargin=0.75 * inch, leftMargin=0.75 * inch,
            topMargin=0.75 * inch,   bottomMargin=0.9 * inch,
        )
        story = []
        S     = self.S

        report_date = insights.get("report_date", datetime.now().strftime("%B %Y"))
        story.append(self._header_block(
            "Labor Market Intelligence Report",
            f"Autonomous AI Analysis  \u00b7  {report_date}",
        ))
        story.append(Spacer(1, 14))

        story.append(KeepTogether([
            self._section_header("Executive Summary"),
            Spacer(1, 6),
        ]))
        for para in insights.get("executive_summary", "").split("\n\n"):
            if para.strip():
                story.append(Paragraph(para.strip(), S["LMBody"]))
        story.append(Spacer(1, 12))

        unemp  = insights.get("unemployment", {})
        wages  = insights.get("wages", {})
        policy = insights.get("policy_structural", {})
        kpis = [
            ("Trend Direction",    unemp.get("trend_direction",  "N/A")),
            ("Headline Signal",    unemp.get("headline",         "N/A")),
            ("Overall Wage Trend", wages.get("overall_wage_trend","N/A")),
            ("Wage Dispersion",    wages.get("wage_dispersion_signal","N/A")),
            ("Structural Shift",   policy.get("structural_shift","N/A")),
            ("Policy Focus",       policy.get("recommended_policy_focus","N/A")),
        ]
        story.append(KeepTogether([
            self._section_header("Key Indicators"),
            Spacer(1, 6),
            self._kpi_grid(kpis),
            Spacer(1, 14),
        ]))

        story.append(KeepTogether([
            self._section_header("Unemployment Dynamics"),
            Spacer(1, 6),
            self._kv_table(
                [["Indicator", "Finding"],
                 ["Trend Direction",       unemp.get("trend_direction",        "N/A")],
                 ["Key Finding",           unemp.get("key_finding",            "N/A")],
                 ["Phillips Curve Signal", unemp.get("phillips_curve_signal",  "N/A")]],
                [2.2 * inch, 4.8 * inch],
            ),
            Spacer(1, 10),
        ]))

        risks = unemp.get("risk_factors", [])
        if risks:
            story.append(KeepTogether([
                Paragraph("Identified Risk Factors", S["LMLabel"]),
                Spacer(1, 4),
                self._risks_table(risks),
                Spacer(1, 14),
            ]))

        story.append(KeepTogether([
            self._section_header("Wage Dynamics"),
            Spacer(1, 6),
            self._kv_table(
                [["Indicator", "Finding"],
                 ["Overall Wage Trend",  wages.get("overall_wage_trend",       "N/A")],
                 ["Sector Leaders",      ", ".join(wages.get("sector_leaders", []))],
                 ["Wage Dispersion",     wages.get("wage_dispersion_signal",   "N/A")],
                 ["Real Wage Status",    wages.get("real_wage_assessment",     "N/A")],
                 ["Policy Implication",  wages.get("policy_implication",       "N/A")]],
                [2.2 * inch, 4.8 * inch],
            ),
            Spacer(1, 14),
        ]))

        story.append(KeepTogether([
            self._section_header("Structural Labor Market Analysis"),
            Spacer(1, 6),
            self._kv_table(
                [["Indicator", "Finding"],
                 ["Structural Shift",   policy.get("structural_shift",          "N/A")],
                 ["Automation Risk",    policy.get("automation_risk_signal",     "N/A")],
                 ["High-Demand Roles",  ", ".join(policy.get("high_demand_occupations", []))],
                 ["Recommended Policy", policy.get("recommended_policy_focus",   "N/A")],
                 ["Confidence Level",   policy.get("confidence_level",           "N/A")]],
                [2.2 * inch, 4.8 * inch],
            ),
            Spacer(1, 20),
        ]))

        doc.build(story, onFirstPage=self._page_footer, onLaterPages=self._page_footer)
        print(f"[ReportAgent] PDF saved: {self.path}")


class ReportAgent:
    def generate_reports(self, insights: dict, csv_datasets: dict) -> dict[str, Path]:
        """
        Generate Excel + PDF labor market reports.
        Accepts output from EconomicAnalysisAgent.run_full_analysis().
        Returns {"excel": Path, "pdf": Path}.
        """
        print("\n=== Report Agent: Generating ===")
        ts = datetime.now().strftime("%Y%m")

        xl_path = OUTPUT_DIR / f"labor_market_{ts}.xlsx"
        wb      = Workbook()
        b       = _ExcelBuilder(wb)
        b.build_summary(insights)
        b.build_unemployment(csv_datasets, insights)
        b.build_wages(csv_datasets, insights)
        b.build_onet(insights)
        b.build_risks(insights)
        wb.save(xl_path)
        print(f"[ReportAgent] Excel saved: {xl_path}")

        pdf_path = OUTPUT_DIR / f"labor_market_{ts}.pdf"
        _PDFBuilder(pdf_path).build(insights, csv_datasets)

        return {"excel": xl_path, "pdf": pdf_path}